"""
患者数据导出模块
提供 Excel 格式的患者列表导出功能
"""
from io import BytesIO
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side


# Excel 表头配置
PATIENT_EXPORT_COLUMNS = [
    ('姓名', 'name', 15),
    ('病案号', 'case_number', 18),
    ('性别', 'gender', 8),
    ('出生日期', 'birth_date', 14),
    ('建档日期', 'created_at', 14),
    ('状态', 'status', 10),
    ('临床诊断', 'clinical_diagnosis', 40),
]

GENDER_MAP = {
    'Male': '男',
    'Female': '女',
}


def export_patients_to_excel(queryset):
    """
    将患者 queryset 导出为 Excel 文件流

    Args:
        queryset: Patient QuerySet（已过滤/排序）

    Returns:
        BytesIO: 包含 Excel 文件内容的字节流
    """
    wb = Workbook()
    ws = wb.active
    ws.title = '患者列表'

    # ---------- 样式 ----------
    header_font = Font(name='微软雅黑', bold=True, size=11, color='FFFFFF')
    header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
    header_alignment = Alignment(horizontal='center', vertical='center')
    cell_font = Font(name='微软雅黑', size=10)
    cell_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin'),
    )

    # ---------- 表头 ----------
    for col_idx, (label, _, width) in enumerate(PATIENT_EXPORT_COLUMNS, start=1):
        cell = ws.cell(row=1, column=col_idx, value=label)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thin_border
        ws.column_dimensions[cell.column_letter].width = width

    # ---------- 数据行 ----------
    for row_idx, patient in enumerate(queryset, start=2):
        for col_idx, (_, field, _) in enumerate(PATIENT_EXPORT_COLUMNS, start=1):
            value = getattr(patient, field, '')

            # 格式化
            if field == 'gender':
                value = GENDER_MAP.get(value, value)
            elif field in ('birth_date', 'created_at') and value:
                if hasattr(value, 'strftime'):
                    value = value.strftime('%Y-%m-%d')

            cell = ws.cell(row=row_idx, column=col_idx, value=value or '')
            cell.font = cell_font
            cell.alignment = cell_alignment
            cell.border = thin_border

    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return output
